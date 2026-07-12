# -*- coding: utf-8 -*-
"""表格：Excel/CSV/TSV 读预览、筛选、排序、统计、转 JSON/CSV/Markdown、列工作表。扁平参数。"""

from __future__ import annotations

import csv
import io
import json
import re
import sys
from pathlib import Path
from typing import Any, List, Optional, Tuple

import agent_common as ac


def _load_csv(path: Path, *, delimiter: str = ",") -> List[dict]:
    enc = "utf-8-sig"
    with open(path, "r", encoding=enc) as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        return list(reader)


def _load_excel(path: Path, sheet: Optional[str]) -> List[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        rows = []
        headers: List[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else f"col{j}" for j, c in enumerate(row)]
                continue
            row_dict = {}
            for j, val in enumerate(row):
                col_name = headers[j] if j < len(headers) else f"col{j}"
                row_dict[col_name] = val
            rows.append(row_dict)
        return rows
    finally:
        wb.close()


def _detect_format(path: Path) -> str:
    s = path.suffix.lower()
    if s in (".xlsx",):
        return "xlsx"
    if s == ".csv":
        return "csv"
    if s in (".tsv", ".tab"):
        return "tsv"
    return "unknown"


def _load_table(path: Path, sheet: Optional[str]) -> Tuple[str, List[dict]]:
    fmt = _detect_format(path)
    if fmt == "xlsx":
        return fmt, _load_excel(path, sheet)
    if fmt in ("csv", "tsv"):
        delim = "\t" if fmt == "tsv" else ","
        return fmt, _load_csv(path, delimiter=delim)
    raise ValueError(f"不支持的文件格式: {path.suffix}（支持 .xlsx / .csv / .tsv）")


def _do_preview(rows: List[dict], limit: int) -> dict:
    if not rows:
        return {"total_rows": 0, "columns": [], "preview": [], "limit": limit}
    columns = list(rows[0].keys())
    return {"total_rows": len(rows), "columns": columns, "preview": rows[:limit], "limit": limit}


def _do_filter(rows: List[dict], col: str, val: str, regex: bool) -> List[dict]:
    if regex:
        pat = re.compile(val, re.IGNORECASE)
        return [r for r in rows if str(r.get(col, "")).strip() and pat.search(str(r.get(col, "")))]
    return [r for r in rows if str(r.get(col, "")).strip().lower() == val.lower()]


def _do_sort(rows: List[dict], col: str, desc: bool) -> List[dict]:
    return sorted(
        rows,
        key=lambda r: (str(r.get(col, "")) if r.get(col) is not None else ""),
        reverse=desc,
    )


def _do_stats(rows: List[dict], col: str) -> dict:
    values = []
    for r in rows:
        v = r.get(col)
        if v is not None and v != "":
            try:
                values.append(float(v))
            except (ValueError, TypeError):
                pass
    stats: dict = {"column": col, "total_rows": len(rows), "non_empty": len(values)}
    if values:
        stats["min"] = min(values)
        stats["max"] = max(values)
        stats["sum"] = round(sum(values), 6)
        stats["avg"] = round(sum(values) / len(values), 6)
    return stats


def _do_to_json(rows: List[dict]) -> str:
    return json.dumps(rows, ensure_ascii=False, indent=2, default=str)


def _do_to_csv(rows: List[dict]) -> str:
    if not rows:
        return ""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def _do_to_md(rows: List[dict], limit: int) -> str:
    if not rows:
        return "(空表)"
    cols = list(rows[0].keys())
    lines = ["|" + "|".join(cols) + "|", "|" + "|".join("---" for _ in cols) + "|"]
    data = rows[:limit] if limit else rows
    for r in data:
        row_vals = []
        for c in cols:
            v = r.get(c)
            s = "" if v is None else str(v).replace("|", "\\|")
            row_vals.append(s)
        lines.append("|" + "|".join(row_vals) + "|")
    return "\n".join(lines)


def agent_main(
    *,
    action: str,
    path: str = "",
    sheet: Optional[str] = None,
    limit: int = 20,
    filter_col: Optional[str] = None,
    filter_val: Optional[str] = None,
    regex: bool = False,
    sort_col: Optional[str] = None,
    sort_desc: bool = False,
    col: Optional[str] = None,
    **_kwargs: Any,
) -> dict:
    if _kwargs.get("source"):
        return ac.err(
            ValueError("data_table 使用 path 指定表格文件，勿传已废弃的 source 参数")
        )
    act = str(action or "").strip().lower().replace("-", "_")
    src_arg = str(path or "").strip()
    if not src_arg:
        return ac.err(ValueError("缺少 path（表格文件路径）"))
    try:
        src = ac.resolve_path(src_arg, allow_outside_workspace=True)
        if not src.is_file():
            return ac.err(FileNotFoundError(f"文件不存在: {src}"))

        fmt, rows = _load_table(src, sheet)

        if act == "sheets":
            if fmt != "xlsx":
                return ac.ok({"sheets": [src.name]})
            import openpyxl

            wb = openpyxl.load_workbook(src, read_only=True)
            try:
                names = wb.sheetnames
            finally:
                wb.close()
            return ac.ok({"sheets": names})

        if act == "preview":
            return ac.ok(_do_preview(rows, limit))
        if act == "filter":
            if not filter_col or filter_val is None:
                return ac.err(ValueError("filter 需要 filter_col 与 filter_val"))
            filtered = _do_filter(rows, filter_col, str(filter_val), regex)
            data = _do_preview(filtered, limit)
            data["filtered_rows"] = len(filtered)
            return ac.ok(data)
        if act == "sort":
            if not sort_col:
                return ac.err(ValueError("sort 需要 sort_col"))
            sorted_rows = _do_sort(rows, sort_col, sort_desc)
            return ac.ok(_do_preview(sorted_rows, limit))
        if act == "stats":
            if not col:
                return ac.err(ValueError("stats 需要 col"))
            return ac.ok(_do_stats(rows, col))
        if act == "to_json":
            return ac.ok({"content": _do_to_json(rows), "format": "json"})
        if act == "to_csv":
            return ac.ok({"content": _do_to_csv(rows), "format": "csv"})
        if act == "to_md":
            return ac.ok({"content": _do_to_md(rows, limit), "format": "markdown"})
        return ac.err(ValueError(f"未知 action: {action}"))
    except Exception as e:
        return ac.err(e)




